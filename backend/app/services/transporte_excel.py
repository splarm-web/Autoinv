"""Parser del Excel de facturas de transporte (formato 'Alfredo').

Replica la lógica de leer_excel() del script generador_facturas.py original:
  - Columnas: Fecha | Viaje | Kilos(toneladas) | Precio   (desde fila 2)
  - El total por línea se calcula = toneladas * precio
  - Kilos mostrados = toneladas * 1000
  - Se corta al encontrar celda vacía o una palabra clave de resumen
  - Debajo puede haber etiquetas opcionales: CABEZA, CISTERNA, NIF, C/C
"""

import io
import re
from datetime import datetime
from typing import Optional

from openpyxl import load_workbook

_STOP_WORDS = ["LAVADOS", "CHEQUE", "BASE", "IVA", "TOTAL",
               "CABEZA", "CISTERNA", "NIF", "C/C"]


def _clean_viaje(nombre: str) -> str:
    """Elimina códigos internos (PV25-1234, OT12345678, números largos)."""
    nombre = re.sub(r"\s+PV\d+-\d+", "", nombre)
    nombre = re.sub(r"\s+OT\d+", "", nombre)
    nombre = re.sub(r"\s+\d{10,}", "", nombre)
    return nombre.strip()


def _to_number(value) -> float:
    """Convierte un valor (string o número) a float, tolerando formato español."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        v = value.strip().replace("€", "").replace(" ", "")
        v = v.replace(".", "").replace(",", ".")
        try:
            return float(v)
        except ValueError:
            return 0.0
    return 0.0


def _fecha_to_iso(value) -> Optional[str]:
    """Normaliza la fecha a 'YYYY-MM-DD' si es posible; si no, devuelve el texto."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    s = str(value).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return s


def parse(content: bytes) -> dict:
    """Lee el Excel y devuelve un dict con líneas, extras y totales calculados.

    Estructura devuelta:
      {
        "viajes": [{fecha, viaje, toneladas, kilos, precio, total}, ...],
        "cabeza": str, "cisterna": str, "nif": str, "ccc": str,
        "base": float, "irpf": float, "iva": float, "total": float
      }
    """
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    viajes = []
    fila = 2
    while True:
        fecha_cell = ws.cell(row=fila, column=1).value
        if fecha_cell is None or fecha_cell == "":
            break
        if isinstance(fecha_cell, str):
            if any(w in fecha_cell.upper() for w in _STOP_WORDS):
                break
        viaje_cell = ws.cell(row=fila, column=2).value
        if viaje_cell is None or viaje_cell == "":
            break

        toneladas = _to_number(ws.cell(row=fila, column=3).value)
        precio = _to_number(ws.cell(row=fila, column=4).value)
        viajes.append({
            "fecha": _fecha_to_iso(fecha_cell),
            "viaje": _clean_viaje(str(viaje_cell)),
            "toneladas": toneladas,
            "kilos": toneladas * 1000,
            "precio": precio,
            "total": round(toneladas * precio, 2),
        })
        fila += 1

    # Etiquetas adicionales debajo de la tabla
    extras = {"cabeza": "", "cisterna": "", "nif": "", "ccc": ""}
    label_map = {"CABEZA": "cabeza", "CISTERNA": "cisterna",
                 "NIF": "nif", "C/C": "ccc"}
    for row in ws.iter_rows(min_row=fila, max_row=ws.max_row, min_col=1, max_col=3):
        label_cell, value_cell = row[0], row[1]
        if label_cell.value:
            label = str(label_cell.value).strip().upper()
            if label in label_map:
                extras[label_map[label]] = (
                    str(value_cell.value) if value_cell.value is not None else ""
                )

    wb.close()

    base = round(sum(v["total"] for v in viajes), 2)
    return {
        "viajes": viajes,
        **extras,
        **_totales(base),
    }


def _totales(base: float) -> dict:
    """IRPF 1%, IVA 21%, total = base - irpf + iva."""
    irpf = round(base * 0.01, 2)
    iva = round(base * 0.21, 2)
    return {
        "base": base,
        "irpf": irpf,
        "iva": iva,
        "total": round(base - irpf + iva, 2),
    }
